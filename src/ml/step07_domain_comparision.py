import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.manifold import TSNE
from utils.utils_ml import load_class_mapping, load_X_y_groups
import openpyxl
from openpyxl.styles import Font

def run_domain_comparison(output_dir):
    print("\n--- [Starting Domain Comparison] ---")
    os.makedirs(output_dir, exist_ok=True)

    print("Calculating F1-score degradation across domains...")
    
    f1_ansys_path = os.path.join(output_dir, "ansys_f1.npy")
    f1_cwru_path = os.path.join(output_dir, "cwru_f1.npy")
    
    if not (os.path.exists(f1_ansys_path) and os.path.exists(f1_cwru_path)):
        raise FileNotFoundError("Could not find ansys_f1.npy or cwru_f1.npy in output/. Please run training and CWRU evaluation first.")
        
    f1_ansys = np.load(f1_ansys_path)
    f1_cwru = np.load(f1_cwru_path)
    
    f1_drop = f1_cwru - f1_ansys
    
    unique_classes, display_labels = load_class_mapping()
    
    df = pd.DataFrame({
        "Fault Class": display_labels,
        "F1 CWRU": f1_cwru,
        "F1 Ansys": f1_ansys,
        "F1 Drop": f1_drop
    })
    
    df.loc[len(df)] = ["Average", f1_cwru.mean(), f1_ansys.mean(), f1_drop.mean()]
    df[["F1 CWRU", "F1 Ansys", "F1 Drop"]] = df[["F1 CWRU", "F1 Ansys", "F1 Drop"]].round(4)
    
    table_path = os.path.join(output_dir, "f1_comparison_table.xlsx")
    df.to_excel(table_path, index=False)
    
    wb = openpyxl.load_workbook(table_path)
    ws = wb.active
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        
    wb.save(table_path)
    print(f"-> Successfully saved F1 Table to '{table_path}'")
    
    print("Preparing t-SNE plot...")
    
    X_ansys, y_ansys, groups_ansys= load_X_y_groups(prefix='ansys')
    X_cwru, y_cwru, groups_cwru= load_X_y_groups(prefix='cwru')
    combined_features = np.vstack([X_ansys, X_cwru])
    
    tsne = TSNE(n_components=2, perplexity=30, random_state=9, n_jobs=-1)
    X_2d = tsne.fit_transform(combined_features)
    
    X_ansys_2d = X_2d[:len(X_ansys)]
    X_cwru_2d = X_2d[len(X_ansys):]
    
    plt.figure(figsize=(12, 9))

    colors = plt.cm.tab10(np.linspace(0, 1, len(unique_classes)))
    
    for c in unique_classes:
        class_name = display_labels[c]
        color = colors[c]
        
        mask_ansys = (y_ansys == c)
        mask_cwru = (y_cwru == c)
        
        if np.any(mask_ansys):
            plt.scatter(
                X_ansys_2d[mask_ansys, 0], X_ansys_2d[mask_ansys, 1], 
                facecolors='none', edgecolors=color, linewidths=1.5,
                alpha=0.6, s=25, label=f"Sim - {class_name}"
            )
            
        if np.any(mask_cwru):
            plt.scatter(
                X_cwru_2d[mask_cwru, 0], X_cwru_2d[mask_cwru, 1], 
                color=color, alpha=0.6, s=25, label=f"Real - {class_name}"
            )
            
    plt.title("t-SNE Space Alignment: ANSYS (Hollow) vs CWRU (Solid)", fontsize=14, fontweight='bold')
    
    plt.legend(bbox_to_anchor=(1.04, 1), loc="upper left", frameon=True, fontsize=12)
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.tight_layout()
    
    tsne_path = os.path.join(output_dir, "domain_tsne_plot.png")
    plt.savefig(tsne_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"-> Successfully saved Multi-Class t-SNE Plot to '{tsne_path}'")

if __name__ == "__main__":
    run_domain_comparison()